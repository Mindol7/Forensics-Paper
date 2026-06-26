from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from Core.normalizer import NormalizedPaper


EXPORT_COLUMNS: list[str] = [
    "source",
    "source_id",
    "category",
    "keywords",
    "title",
    "title_kor",
    "title_eng",
    "doi",
    "url",
    "abstract",
    "paper_keywords",
    "authors",
    "journal_id",
    "institution_id",
    "issue_id",
    "issn",
    "subject_code",
    "registered_at",
    "publication_year",
    "relevance_score",
    "summary",
]

PREFERRED_WIDTHS: dict[str, int] = {
    "source": 10,
    "source_id": 18,
    "category": 28,
    "keywords": 30,
    "title": 48,
    "title_kor": 48,
    "title_eng": 48,
    "doi": 28,
    "url": 36,
    "abstract": 72,
    "paper_keywords": 34,
    "authors": 30,
    "journal_id": 26,
    "institution_id": 26,
    "issue_id": 14,
    "issn": 16,
    "subject_code": 26,
    "registered_at": 16,
    "publication_year": 14,
    "relevance_score": 14,
    "summary": 54,
}

WRAPPED_COLUMNS: set[str] = {
    "category",
    "keywords",
    "title",
    "title_kor",
    "title_eng",
    "abstract",
    "paper_keywords",
    "authors",
    "journal_id",
    "institution_id",
    "subject_code",
    "summary",
}

INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")


def _join(values: Iterable[str]) -> str:
    return "; ".join(value for value in values if value)


def _flatten_paper(paper: NormalizedPaper) -> dict[str, object]:
    return {
        "source": paper.source,
        "source_id": paper.source_id,
        "category": _join(paper.categories),
        "keywords": _join(paper.matched_keywords),
        "title": paper.title,
        "title_kor": paper.title_kor,
        "title_eng": paper.title_eng,
        "doi": paper.doi,
        "url": paper.url,
        "abstract": paper.abstract,
        "paper_keywords": _join(paper.keywords),
        "authors": _join(paper.authors),
        "journal_id": paper.journal_id,
        "institution_id": paper.institution_id,
        "issue_id": paper.issue_id,
        "issn": paper.issn,
        "subject_code": paper.subject_code,
        "registered_at": paper.registered_at,
        "publication_year": paper.publication_year,
        "relevance_score": paper.relevance_score,
        "summary": paper.summary,
    }


def export_papers_to_excel(papers: Iterable[NormalizedPaper], output_path: Path) -> Path:
    paper_list = list(papers)
    rows = [_flatten_paper(paper) for paper in paper_list]
    output_path.parent.mkdir(parents=True, exist_ok=True)

    dataframe = pd.DataFrame(rows, columns=EXPORT_COLUMNS)

    category_rows: dict[str, list[dict[str, object]]] = {}
    for paper, row in zip(paper_list, rows, strict=False):
        categories = paper.categories or ["미분류"]
        for category in categories:
            category_rows.setdefault(category, []).append(row)

    used_sheet_names = {"전체"}
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="전체")
        for category, sheet_rows in category_rows.items():
            sheet_name = _safe_sheet_name(category, used_sheet_names)
            category_frame = pd.DataFrame(sheet_rows, columns=EXPORT_COLUMNS)
            category_frame.to_excel(writer, index=False, sheet_name=sheet_name)

    _style_workbook(output_path)
    return output_path


def _safe_sheet_name(value: str, used_sheet_names: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS_RE.sub(" ", value).strip() or "미분류"
    cleaned = " ".join(cleaned.split())
    base = cleaned[:31] or "미분류"
    name = base
    counter = 2
    while name in used_sheet_names:
        suffix = f" {counter}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_sheet_names.add(name)
    return name


def _style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        _style_worksheet(worksheet)
    workbook.active = 0
    workbook.save(path)


def _style_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        )

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="D9E2F3")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers: dict[int, str] = {}
    for cell in worksheet[1]:
        header = "" if cell.value is None else str(cell.value)
        headers[cell.column] = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = cell_border
    worksheet.row_dimensions[1].height = 24

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            header = headers.get(cell.column, "")
            wrap_text = header in WRAPPED_COLUMNS
            cell.alignment = Alignment(vertical="top", wrap_text=wrap_text)
            cell.border = cell_border

    for index, header in headers.items():
        column_letter = get_column_letter(index)
        preferred_width = PREFERRED_WIDTHS.get(header)
        if preferred_width is not None:
            worksheet.column_dimensions[column_letter].width = preferred_width
            continue

        max_length = len(header)
        for cell in worksheet[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
